from openpyxl import Workbook
from openpyxl.styles import Font, Color, Border, Side, PatternFill, GradientFill, Alignment
import sys

def gera_colunas_excel(limite):
    abc = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    abc_len = len(abc)
    lista = []
    i = 0
    ii = -1
    while(int(limite) > int(len(lista))):
        if len(lista) >= abc_len:
            ii += 1
            j = 0
            
            while(int(limite) > int(len(lista))):
                lista.append(abc[ii] + abc[j])
                j += 1
                if j == abc_len:
                    ii += 1
                    j = 0
        else:
            lista.append(abc[i])
            i += 1

    return lista

def gera_relatorio(titulo="Titulo Excel", localSalvamento = "", nomeArquivo="Relatorio.xlsx",colunas = [], dados = []):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    
    colunas_excel = gera_colunas_excel(len(colunas))

    linha = 1
    index = 0
    for coluna in colunas_excel:
        colName = "{}{}".format(coluna,linha)
        ws[colName] = colunas[index].replace('_',' ')
        ws[colName].font = Font(color="FF0000", italic=True)
        ws[colName].fill = PatternFill("solid", fgColor="2C222B")
        ws[colName].alignment = Alignment(horizontal="center", vertical="center")
        index += 1

    linha = 2
    for dado in dados:
        index = 0
        for coluna in colunas:
            colName = "{}{}".format(colunas_excel[index],linha)
            ws[colName] = dado[coluna]
            index += 1
        linha += 1
    
    wb.save("{}/{}.xlsx".format(localSalvamento,nomeArquivo))
    return True